import os
import subprocess
import sys
from copy import copy
import PySimpleGUI as sg
import csv

from openpyxl.formatting.rule import CellIsRule
from openpyxl.formula.translate import Translator
from openpyxl import load_workbook
from datetime import date, datetime

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


sg.theme('DarkAmber')

excel_file_path = sg.popup_get_file("Please select the excel file to update")

wb = load_workbook(excel_file_path)

current_date = date.today()

wb.save(f"budget_backup_{current_date.strftime('%b-%d-%Y')}.xlsx")

ws = wb.active

layout = [
    [sg.Text("Enter your values here")]
]

input_dict = {}

first_col = ws["A"]

add_wealthscape = False
cell_counter = 0
for cell in first_col:
    cell_counter += 1
    temp_str = cell.value
    if "*" in str(temp_str):
        temp_str = temp_str.replace("*", "")
        input_dict[cell_counter] = temp_str
        layout.append([sg.Text(temp_str), sg.InputText()])
    elif "Wealthscape" in str(temp_str):
        add_wealthscape = True

if add_wealthscape:
    layout.append([sg.Text("Wealthscape File"), sg.Button("Open File")])

layout.append([sg.Button("Ok"), sg.Button("Cancel")])

window = sg.Window('Program', layout)

wealthscape_file_path = None
while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'Cancel':
        print("User cancelled operation")
        window.close()
        sys.exit()
    elif event == "Ok":
        index = 0
        for num in input_dict.keys():
            if values[index] == '':
                values[index] = 0
            values[index] = float(values[index])
            input_dict.update({num: values[index]})
            index += 1
        break
    elif event == "Open File":
        wealthscape_file_path = sg.popup_get_file('Please upload wealthscape file')

mkt_values = []
window.close()
if wealthscape_file_path is not None:
    with open(wealthscape_file_path) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            if len(row) >= 20 and row[20] != "Recent Market Value":
                temp_str = row[20]
                temp_str = temp_str.replace('$', '')
                temp_str = temp_str.replace(',', '')
                mkt_values.append(float(temp_str))

bal_col = None
bal_col_num = None
avg_col = None
avg_col_num = None
col_counter = 0


for col in ws.iter_cols():
    if col[3].value == "Balance":
        bal_col = col
        bal_col_num = col_counter
    elif col[3].value == "Average":
        avg_col = col
        avg_col_num = col_counter
    col_counter += 1

ws.insert_cols(bal_col_num + 3)

bal_col_letter = get_column_letter(bal_col_num + 3)

purple_font = Font(color="7030A0", bold=True)
red_font = Font(color="FF0000", bold=True)
green_font = Font(color="00B050", bold=True)

is_stock = False
cell_counter = 0
mkt_index = 0
for cell in bal_col:
    new_cell = ws[bal_col_letter][cell_counter]
    if cell_counter + 1 in input_dict.keys() and input_dict[cell_counter + 1] != 0:
        new_cell.value = input_dict[cell_counter + 1]
    elif is_stock and mkt_index < len(mkt_values):
        new_cell.value = mkt_values[mkt_index]
        mkt_index += 1
    else:
        new_cell.value = cell.value
    if cell.has_style:
        new_cell.font = copy(cell.font)
        new_cell.border = copy(cell.border)
        new_cell.fill = copy(cell.fill)
        new_cell.number_format = copy(cell.number_format)
        new_cell.protection = copy(cell.protection)
        new_cell.alignment = copy(cell.alignment)
    cell_counter += 1
    if cell.value is not None and is_stock:
        prev_mkt_col_letter = get_column_letter(bal_col_num + 1)
        ws.conditional_formatting.add(bal_col_letter + str(cell_counter),
                                      CellIsRule(operator='>=',
                                                 formula=[f"${prev_mkt_col_letter}{cell_counter}"],
                                                 stopIfTrue=True,
                                                 font=green_font))
        ws.conditional_formatting.add(bal_col_letter + str(cell_counter),
                                      CellIsRule(operator='<',
                                                 formula=[f"${prev_mkt_col_letter}{cell_counter}"],
                                                 stopIfTrue=True,
                                                 font=red_font))
    elif cell.value is not None and cell.value != "Balance" and cell.value != "Mkt Value":
        ws.conditional_formatting.add(bal_col_letter + str(cell_counter),
                                      CellIsRule(operator='between',
                                                 formula=[f"$W{cell_counter}", f"$X{cell_counter}-.01"],
                                                 stopIfTrue=True,
                                                 font=green_font))
        ws.conditional_formatting.add(bal_col_letter + str(cell_counter),
                                      CellIsRule(operator='>=',
                                                 formula=[f"$X{cell_counter}"],
                                                 stopIfTrue=True,
                                                 font=purple_font))
        ws.conditional_formatting.add(bal_col_letter + str(cell_counter),
                                      CellIsRule(operator='<',
                                                 formula=[f"$W{cell_counter}"],
                                                 stopIfTrue=True,
                                                 font=red_font))
    if cell.data_type == "f":
        origin_letter = get_column_letter(bal_col_num + 1)
        ws[bal_col_letter + str(cell_counter)] = \
            Translator(cell.value, origin=origin_letter + str(cell_counter)) \
                .translate_formula(bal_col_letter + str(cell_counter))
    if cell.value == "Mkt Value":
        is_stock = True

col_counter = 0
sum_cols = []
for col in ws.iter_cols():
    if isinstance(col[3].value, datetime):
        col_letter = get_column_letter(col_counter+1)
        sum_cols.append(col_letter)
    col_counter += 1

average_col = ws[get_column_letter(avg_col_num + 2)]
cell_counter = 0
for cell in average_col:
    cell_counter += 1
    if cell.value is not None and cell.value != 2022 and cell.value != "Average" and "AVERAGE" in cell.value:
        cell.value = f"=AVERAGE(Y{cell_counter}:{bal_col_letter}{cell_counter})"
    elif cell.value is not None and cell.value != 2022 and cell.value != "Average":
        temp_str = "=("
        add_counter = 0
        for add in sum_cols:
            temp_str += f"{add}{cell_counter}"
            if add_counter < len(sum_cols) - 1:
                temp_str += "+"
                add_counter += 1
        temp_str += ")"
        if "26" in cell.value:
            temp_str += "/26"
        cell.value = temp_str

next_letter = get_column_letter(bal_col_num + 4)
next_col = ws[next_letter]
for x in range(len(next_col)):
    if ws[next_letter + str(x + 1)].data_type == "f":
        origin_col_letter = get_column_letter(bal_col_num + 3)
        origin_str = origin_col_letter + str(cell_counter)

        translate_str = next_letter + str(cell_counter)
        ws[next_letter + str(x + 1)] = Translator(ws[next_letter + str(x + 1)].value,
                                                  origin=origin_str).translate_formula(translate_str)

wb.save(excel_file_path)

sg.popup("Process finished!")

os.startfile(excel_file_path, 'open')
